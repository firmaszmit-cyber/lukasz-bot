from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_xlsx(klient: str, prace: list[dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = klient[:30] if klient else "Kosztorys"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    total_fill = PatternFill("solid", fgColor="C6EFCE")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="888888")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    money_fmt = '#,##0.00'

    ws["A1"] = "KOSZTORYS — SZMIT Remonty"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 22

    ws["A2"] = f"Klient: {klient}"
    ws["A2"].font = bold
    ws.merge_cells("A2:F2")
    ws["A3"] = f"Data: {date.today().strftime('%d.%m.%Y')}"
    ws.merge_cells("A3:F3")

    headers = ["Lp.", "Opis pozycji", "Ilość", "Jedn.", "Stawka (zł)", "Wartość (zł)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[5].height = 22

    r = 6
    total_min = 0.0
    total_max = 0.0

    for i, p in enumerate(prace, 1):
        stawka_min = p["stawka_min"]
        stawka_max = p.get("stawka_max", stawka_min)
        ilosc = p["ilosc"]
        stawka = (stawka_min + stawka_max) / 2 if stawka_max != stawka_min else stawka_min
        wartosc = round(ilosc * stawka, 2)
        total_min += ilosc * stawka_min
        total_max += ilosc * stawka_max

        ws.cell(row=r, column=1, value=i).alignment = center
        ws.cell(row=r, column=2, value=p["nazwa"]).alignment = left
        ws.cell(row=r, column=3, value=ilosc).alignment = center
        ws.cell(row=r, column=4, value=p["jednostka"]).alignment = center
        ws.cell(row=r, column=5, value=stawka).alignment = right
        ws.cell(row=r, column=6, value=wartosc).alignment = right
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=5).number_format = money_fmt
        ws.cell(row=r, column=6).number_format = money_fmt
        r += 1

    r += 1
    total = (total_min + total_max) / 2 if total_max != total_min else total_min
    label = ws.cell(row=r, column=1, value="SUMA ROBOCIZNA")
    label.font = Font(bold=True, size=12)
    label.alignment = right
    val = ws.cell(row=r, column=6, value=round(total, 2))
    val.font = Font(bold=True, size=12)
    val.alignment = right
    val.number_format = money_fmt
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = total_fill
        ws.cell(row=r, column=col).border = border
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 24

    r += 2
    ws.cell(row=r, column=1, value="Uwagi:").font = bold
    r += 1
    for line in [
        "• Ceny dotyczą robocizny. Materiały budowlane po stronie klienta.",
        "• Gwarancja 24 miesiące.",
        "• Stawki zgodne z cennikiem SZMIT Remonty.",
    ]:
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    r += 1
    footer = ws.cell(row=r, column=1,
                     value="SZMIT Remonty — Zawsze na czas  |  +48 692 238 159  |  firmaszmit@gmail.com  |  Kraków")
    footer.font = Font(italic=True, size=9, color="666666")
    footer.alignment = left
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
